import sqlite3
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==========================================
# CONFIGURACAO
# ==========================================
DB_PECAS = os.path.join("automacao_pecas", "oab_questoes.db")
DB_DISCURSIVAS = os.path.join("automacao_discursivas", "oab_discursivas.db")
ARQUIVO_SAIDA = "questoes_oab_2fase.xlsx"

# Cores do tema
COR_HEADER = "1F2937"       # Cinza escuro
COR_HEADER_FONT = "FFFFFF"  # Branco
COR_SEPARADOR = "DBEAFE"    # Azul claro (separador de exame)
COR_PECA = "FEF3C7"         # Amarelo claro (linha de peca)
COR_ZEBRA_1 = "FFFFFF"      # Branco
COR_ZEBRA_2 = "F8FAFC"      # Cinza bem claro

# ==========================================
# FUNCOES DE ESTILO
# ==========================================
def estilo_header():
    return {
        "font": Font(name="Segoe UI", bold=True, color=COR_HEADER_FONT, size=11),
        "fill": PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(bottom=Side(style="thin", color="4B5563"))
    }

def estilo_separador():
    return {
        "font": Font(name="Segoe UI", bold=True, size=10, color="1E40AF"),
        "fill": PatternFill(start_color=COR_SEPARADOR, end_color=COR_SEPARADOR, fill_type="solid"),
        "alignment": Alignment(vertical="center"),
        "border": Border(bottom=Side(style="thin", color="93C5FD"))
    }

def estilo_celula(tipo="discursiva", zebra=False):
    if tipo == "peca":
        fill_color = COR_PECA
    else:
        fill_color = COR_ZEBRA_2 if zebra else COR_ZEBRA_1
    return {
        "font": Font(name="Segoe UI", size=10),
        "fill": PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid"),
        "alignment": Alignment(vertical="top", wrap_text=True),
        "border": Border(bottom=Side(style="hair", color="E5E7EB"))
    }

def aplicar_estilo(cell, estilo):
    for attr, value in estilo.items():
        setattr(cell, attr, value)

# ==========================================
# LEITURA DO BANCO DE DADOS
# ==========================================
def ler_pecas():
    if not os.path.exists(DB_PECAS):
        print(f"[AVISO] Banco de pecas nao encontrado: {DB_PECAS}")
        return []
    conn = sqlite3.connect(DB_PECAS)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT area_direito, exame_nome, enunciado, resposta_padrao, distribuicao_pontos
        FROM pecas_fgv
        ORDER BY area_direito, exame_nome
    """)
    dados = cursor.fetchall()
    conn.close()
    return dados

def ler_discursivas():
    if not os.path.exists(DB_DISCURSIVAS):
        print(f"[AVISO] Banco de discursivas nao encontrado: {DB_DISCURSIVAS}")
        return []
    conn = sqlite3.connect(DB_DISCURSIVAS)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT area_direito, exame_nome, enunciado, resposta_a, resposta_b, pontuacao_a, pontuacao_b, distribuicao_pontos
        FROM discursivas_fgv
        ORDER BY area_direito, exame_nome
    """)
    dados = cursor.fetchall()
    conn.close()
    return dados

# ==========================================
# CASAR QUESTOES POR EXAME
# ==========================================
def organizar_por_exame(pecas, discursivas):
    """
    Agrupa pecas e discursivas pelo exame (area + exame_nome).
    Retorna uma lista ordenada de blocos, cada um com:
      { "area": ..., "exame": ..., "peca": ..., "discursivas": [...] }
    """
    exames = {}
    
    # Indexar pecas por chave (area + exame)
    for area, exame, enunciado, resposta, dist_pontos in pecas:
        chave = (area, exame)
        if chave not in exames:
            exames[chave] = {"area": area, "exame": exame, "pecas": [], "discursivas": []}
        exames[chave]["pecas"].append((enunciado, resposta, dist_pontos or ""))
    
    # Indexar discursivas pela mesma chave
    for area, exame, enunciado, resp_a, resp_b, pont_a, pont_b, dist_pontos in discursivas:
        chave = (area, exame)
        if chave not in exames:
            exames[chave] = {"area": area, "exame": exame, "pecas": [], "discursivas": []}
        exames[chave]["discursivas"].append((enunciado, resp_a, resp_b, pont_a or "", pont_b or "", dist_pontos or ""))
    
    # Ordenar por area e exame
    ordenados = sorted(exames.values(), key=lambda x: (x["area"], x["exame"]))
    return ordenados

# ==========================================
# CRIACAO DA PLANILHA UNIFICADA
# ==========================================
def criar_planilha_unificada(wb, blocos):
    ws = wb.active
    ws.title = "Prova OAB 2a Fase"
    
    # Cabecalhos
    headers = ["Area", "Exame", "Tipo", "Enunciado", "Padrao de Resposta", "Resposta A", "Resposta B", "Pontuacao A", "Pontuacao B", "Distribuicao de Pontos"]
    larguras = [22, 35, 20, 80, 70, 55, 55, 18, 18, 70]
    
    header_style = estilo_header()
    for col, (header, largura) in enumerate(zip(headers, larguras), 1):
        cell = ws.cell(row=1, column=col, value=header)
        aplicar_estilo(cell, header_style)
        ws.column_dimensions[cell.column_letter].width = largura
    
    ws.row_dimensions[1].height = 30
    
    row_idx = 2
    total_questoes = 0
    
    for bloco in blocos:
        area = bloco["area"]
        exame = bloco["exame"]
        
        # --- Linha separadora do exame ---
        sep_style = estilo_separador()
        sep_text = f"{area} - {exame}"
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col, value=sep_text if col == 1 else "")
            aplicar_estilo(cell, sep_style)
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        ws.row_dimensions[row_idx].height = 25
        row_idx += 1
        
        # --- Pecas (todas) ---
        for j, (enunciado, resposta, dist_pontos) in enumerate(bloco["pecas"]):
            valores = [area, exame, f"Peca Profissional {j+1}" if len(bloco["pecas"]) > 1 else "Peca Profissional", enunciado, resposta, "", "", "", "", dist_pontos]
            peca_style = estilo_celula(tipo="peca")
            for col, valor in enumerate(valores, 1):
                cell = ws.cell(row=row_idx, column=col, value=valor)
                aplicar_estilo(cell, peca_style)
            row_idx += 1
            total_questoes += 1
        
        # --- Discursivas ---
        for i, (enunciado, resp_a, resp_b, pont_a, pont_b, dist_pontos) in enumerate(bloco["discursivas"]):
            zebra = (i % 2 == 0)
            disc_style = estilo_celula(tipo="discursiva", zebra=zebra)
            valores = [area, exame, f"Questao Discursiva {i+1}", enunciado, "", resp_a, resp_b, pont_a, pont_b, dist_pontos]
            for col, valor in enumerate(valores, 1):
                cell = ws.cell(row=row_idx, column=col, value=valor or "")
                aplicar_estilo(cell, disc_style)
            row_idx += 1
            total_questoes += 1
    
    # Congelar cabecalho
    ws.freeze_panes = "A2"
    
    # Filtro automatico
    if total_questoes > 0:
        ws.auto_filter.ref = f"A1:J{row_idx - 1}"
    
    return total_questoes

# ==========================================
# FLUXO PRINCIPAL
# ==========================================
def exportar():
    print(f"\n{'='*60}")
    print(f"  EXPORTADOR OAB 2a FASE -> EXCEL (UNIFICADO)")
    print(f"{'='*60}")
    
    wb = Workbook()
    
    print("\n[1/4] Lendo banco de pecas...")
    pecas = ler_pecas()
    print(f"      -> {len(pecas)} peca(s) encontrada(s).")
    
    print("[2/4] Lendo banco de discursivas...")
    discursivas = ler_discursivas()
    print(f"      -> {len(discursivas)} discursiva(s) encontrada(s).")
    
    if not pecas and not discursivas:
        print("\n[AVISO] Nenhum dado encontrado nos bancos. Execute os robos primeiro!")
        return
    
    print("[3/4] Organizando por exame e montando planilha...")
    blocos = organizar_por_exame(pecas, discursivas)
    print(f"      -> {len(blocos)} exame(s) identificado(s).")
    
    total = criar_planilha_unificada(wb, blocos)
    
    print(f"[4/4] Salvando '{ARQUIVO_SAIDA}'...")
    wb.save(ARQUIVO_SAIDA)
    
    print(f"\n{'='*60}")
    print(f"  [OK] EXPORTACAO CONCLUIDA!")
    print(f"  Arquivo: {os.path.abspath(ARQUIVO_SAIDA)}")
    print(f"  Exames: {len(blocos)} | Questoes: {total}")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    exportar()
